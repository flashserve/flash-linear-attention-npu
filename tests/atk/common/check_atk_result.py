#!/usr/bin/env python3
"""ATK 测试结果检查器。

解析 ATK 生成的 xlsx 报告，检查 accuracy / determinism / mssanitizer 是否全部通过，
统计 fail 数量。供 run_test_cpu.sh 在每个测试阶段结束后调用。

用法:
  python3 common/check_atk_result.py --type accuracy --output-root ./atk_output --op chunk_bwd_dqkwg
  python3 common/check_atk_result.py --type all --output-root ./atk_output --op chunk_bwd_dqkwg

退出码: 0=全部通过, 1=存在失败, 2=检查器自身错误(如缺依赖)
"""

import argparse
import glob
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[ATK结果检查] 错误: 需要 openpyxl，请 pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# 报告查找
# ---------------------------------------------------------------------------

def _find_xlsx_files(output_root, pattern):
    """返回按修改时间排序的 xlsx 文件列表（旧→新）。"""
    full_pattern = os.path.join(output_root, pattern, "report", "*.xlsx")
    return sorted(glob.glob(full_pattern), key=os.path.getmtime)


def _find_accuracy_reports(output_root, op):
    """accuracy 报告在 cpu_dual_reference/atk_output/atk_<op>_* 下。"""
    return _find_xlsx_files(output_root, f"cpu_dual_reference/atk_output/atk_{op}_*")


def _find_accuracy_gpu_reports(output_root, op):
    """accuracy_gpu 报告在 gpu_dual_reference/atk_output/atk_<op>_* 下。"""
    return _find_xlsx_files(output_root, f"gpu_dual_reference/atk_output/atk_{op}_*")


def _find_root_reports(output_root, op):
    """determinism 和 mssanitizer 报告共享 atk_<op>_* 目录。"""
    return _find_xlsx_files(output_root, f"atk_{op}_*")


# ---------------------------------------------------------------------------
# 报告解析
# ---------------------------------------------------------------------------

def _parse_summary(xlsx_path):
    """解析 xlsx 的 summary sheet，返回 (header_list, data_rows)。

    header_list: 第一行表头列表
    data_rows: 剩余行列表，每行为单元格值列表
    找不到 summary sheet 或文件损坏时返回 (None, None)。
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    except Exception:
        return None, None
    if "summary" not in wb.sheetnames:
        wb.close()
        return None, None
    ws = wb["summary"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None, None
    header = [str(c) if c is not None else "" for c in rows[0]]
    data = [[c for c in r] for r in rows[1:]]
    return header, data


def _is_mssanitizer_report(header):
    """mssanitizer 报告的表头包含 '内存检测' 列。"""
    return any("内存检测" in h for h in header)


def _classify_report(header):
    """根据表头分类报告类型。

    返回 'accuracy' | 'mssanitizer' | 'unknown'
    accuracy 和 determinism 的表头结构相同（含 精度是否达标），无法仅凭表头区分，
    需要调用方根据所在目录判断。
    """
    if _is_mssanitizer_report(header):
        return "mssanitizer"
    if any("精度是否达标" in h or "通过率" in h for h in header):
        return "accuracy"
    return "unknown"


def _extract_summary_row(header, data_rows):
    """从 summary 行中提取关键指标。

    返回 dict: {total, exec_pass, exec_fail, check_pass, check_fail,
               pass_rate,达标}
    各字段不存在时为 None。
    """
    def _col_idx(name_part):
        for i, h in enumerate(header):
            if name_part in h:
                return i
        return None

    total_idx = _col_idx("总用例数")
    exec_pass_idx = _col_idx("执行成功用例个数")
    exec_fail_idx = _col_idx("执行失败用例个数")
    check_pass_idx = _col_idx("通过用例个数")
    err_match_idx = _col_idx("错误信息匹配用例个数")
    rate_idx = _col_idx("通过率")
    mss_rate_idx = _col_idx("内存检测通过率")
    # 收集所有"达标"列索引
    # accuracy: 精度是否达标
    # determinism: 精度是否达标(值为"-") + 确定性计算是否达标(实际结果)
    # mssanitizer: 内存检测是否达标
    pass_indices = [i for i, h in enumerate(header) if "达标" in h]

    # 汇总所有 node 行
    total = exec_pass = exec_fail = check_pass = check_fail = 0
    pass_rate = mss_pass_rate = None
    all_pass = True
    for row in data_rows:
        if not row or all(c is None for c in row):
            continue
        if total_idx is not None and row[total_idx] is not None:
            try:
                total += int(row[total_idx])
            except (ValueError, TypeError):
                pass
        if exec_pass_idx is not None and row[exec_pass_idx] is not None:
            try:
                exec_pass += int(row[exec_pass_idx])
            except (ValueError, TypeError):
                pass
        if exec_fail_idx is not None and row[exec_fail_idx] is not None:
            try:
                exec_fail += int(row[exec_fail_idx])
            except (ValueError, TypeError):
                pass
        if check_pass_idx is not None and row[check_pass_idx] is not None:
            try:
                check_pass += int(row[check_pass_idx])
            except (ValueError, TypeError):
                pass
        if err_match_idx is not None and row[err_match_idx] is not None:
            try:
                check_fail += int(row[err_match_idx])
            except (ValueError, TypeError):
                pass
        if rate_idx is not None and row[rate_idx] is not None:
            pass_rate = row[rate_idx]
        if mss_rate_idx is not None and row[mss_rate_idx] is not None:
            mss_pass_rate = row[mss_rate_idx]
        # 检查所有达标列：若某列值为 "Failed" 则整体失败
        # ("-"/空值 视为不适用，不影响判断)
        for pi in pass_indices:
            if pi < len(row) and row[pi] is not None:
                val = str(row[pi]).strip()
                if val == "Failed":
                    all_pass = False

    # mssanitizer 没有 exec_fail，用 total - (pass count) 估算
    if exec_fail == 0 and total > 0 and check_pass == 0:
        # mssanitizer 情况：total 个用例，达标判断看 "内存检测是否达标"
        exec_pass = total
        check_pass = total if all_pass else 0
        check_fail = 0 if all_pass else total

    return {
        "total": total,
        "exec_pass": exec_pass,
        "exec_fail": exec_fail,
        "check_pass": check_pass,
        "check_fail": check_fail,
        "pass_rate": pass_rate,
        "mss_pass_rate": mss_pass_rate,
        "all_pass": all_pass,
    }


# ---------------------------------------------------------------------------
# 公共 API
# ---------------------------------------------------------------------------

def check_accuracy(output_root, op):
    """检查 accuracy（CPU 标杆）报告。"""
    files = _find_accuracy_reports(output_root, op)
    if not files:
        return {"found": False, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
                "xlsx": None, "detail": "未找到 accuracy 报告"}
    xlsx = files[-1]
    header, data = _parse_summary(xlsx)
    if header is None:
        return {"found": True, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
                "xlsx": xlsx, "detail": "无法解析 summary sheet"}
    info = _extract_summary_row(header, data)
    return {"found": True, **info, "xlsx": xlsx,
            "detail": f"通过率={info['pass_rate']}"}


def check_accuracy_gpu(output_root, op):
    """检查 accuracy_gpu（GPU 标杆）报告。"""
    files = _find_accuracy_gpu_reports(output_root, op)
    if not files:
        return {"found": False, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
                "xlsx": None, "detail": "未找到 accuracy_gpu 报告"}
    xlsx = files[-1]
    header, data = _parse_summary(xlsx)
    if header is None:
        return {"found": True, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
                "xlsx": xlsx, "detail": "无法解析 summary sheet"}
    info = _extract_summary_row(header, data)
    return {"found": True, **info, "xlsx": xlsx,
            "detail": f"通过率={info['pass_rate']}"}


def check_determinism(output_root, op):
    """检查 determinism 报告（与 mssanitizer 共享目录，按表头过滤）。"""
    files = _find_root_reports(output_root, op)
    # 从新到旧找第一个非 mssanitizer 报告
    for f in reversed(files):
        header, _ = _parse_summary(f)
        if header is not None and not _is_mssanitizer_report(header):
            header, data = _parse_summary(f)
            info = _extract_summary_row(header, data)
            return {"found": True, **info, "xlsx": f,
                    "detail": f"通过率={info['pass_rate']}"}
    return {"found": False, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
            "xlsx": None, "detail": "未找到 determinism 报告"}


def check_mssanitizer(output_root, op):
    """检查 mssanitizer 报告。"""
    files = _find_root_reports(output_root, op)
    for f in reversed(files):
        header, _ = _parse_summary(f)
        if header is not None and _is_mssanitizer_report(header):
            header, data = _parse_summary(f)
            info = _extract_summary_row(header, data)
            return {"found": True, **info, "xlsx": f,
                    "detail": f"内存检测通过率={info['mss_pass_rate']}"}
    return {"found": False, "total": 0, "pass": 0, "fail": 0, "all_pass": False,
            "xlsx": None, "detail": "未找到 mssanitizer 报告"}


CHECKERS = {
    "accuracy": check_accuracy,
    "accuracy_cpu": check_accuracy,
    "accuracy_gpu": check_accuracy_gpu,
    "determinism": check_determinism,
    "mssanitizer": check_mssanitizer,
}

TYPE_LABELS = {
    "accuracy": "精度",
    "accuracy_cpu": "CPU精度",
    "accuracy_gpu": "GPU精度",
    "determinism": "确定性",
    "mssanitizer": "内存检测",
}


def main():
    parser = argparse.ArgumentParser(description="ATK 测试结果检查器")
    parser.add_argument("--type", required=True,
                        choices=["accuracy", "determinism", "mssanitizer", "all"],
                        help="检查的测试类型")
    parser.add_argument("--output-root", required=True,
                        help="ATK 输出根目录（如 ./atk_output）")
    parser.add_argument("--op", required=True, help="算子名（如 chunk_bwd_dqkwg）")
    args = parser.parse_args()

    output_root = os.path.abspath(args.output_root)
    types = list(CHECKERS.keys()) if args.type == "all" else [args.type]

    results = {}
    any_fail = False

    for t in types:
        r = CHECKERS[t](output_root, args.op)
        results[t] = r
        label = TYPE_LABELS[t]
        if not r["found"]:
            status_str = "NO_REPORT"
            print(f"[ATK结果检查] {label}: 未找到报告（跳过）")
            continue
        status = "Pass" if r["all_pass"] else "Failed"
        if not r["all_pass"]:
            any_fail = True
        xlsx_name = os.path.basename(r.get("xlsx", "")) if r.get("xlsx") else ""
        # 实际失败数 = 总用例 - 通过用例 (check_fail 是"错误信息匹配"，不等价于未通过)
        actual_fail = r["total"] - r["check_pass"]
        print(f"[ATK结果检查] {label}: {status} "
              f"(总用例={r['total']}, 通过={r['check_pass']}, "
              f"失败={actual_fail}) [{xlsx_name}]")

    if args.type == "all":
        # 仅统计找到报告的测试类型
        found_types = [t for t in types if results[t].get("found")]
        passed_types = sum(1 for t in found_types if results[t].get("all_pass"))
        found_count = len(found_types)
        skipped = len(types) - found_count
        fail_count = found_count - passed_types
        msg = f"[ATK结果检查] 汇总: {passed_types}/{found_count} 通过, {fail_count} 项失败"
        if skipped > 0:
            msg += f", {skipped} 项无报告跳过"
        print(msg)

    sys.exit(1 if any_fail else 0)


if __name__ == "__main__":
    main()
